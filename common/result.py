
import os
import json

from openpyxl import Workbook

from common.globel_veriable import REGISTRATION_DIR_127


def write2excel() -> None:
    """
    Write the metric values to excel.
    :return: None
    """
    columns = ["Case name", "Correlation Metric Value", "Baseline", "BL_Manufacturer", "Follow-up", "FU_Manufacturer"]
    metric_value = os.path.abspath(os.path.join(REGISTRATION_DIR_127, "../..", "127_metric_value.json"))
    meta_data = os.path.abspath(os.path.join(REGISTRATION_DIR_127, "../..", "127_meta_datas.json"))

    with open(metric_value, "r+") as f:
        metric_values = json.load(f)

    with open(meta_data, "r+") as f:
        meta_datas = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "Result"
    for idx, column in enumerate(columns, start=1):
        ws.cell(row=1, column=idx, value=column)

    for idx, (key, value) in enumerate(metric_values.items(), start=2):
        ws.cell(row=idx, column=1, value=key)
        ws.cell(row=idx, column=2, value=value)

    keys = list(meta_datas.keys())
    bl_list = keys[::2]
    fu_list = keys[1::2]

    for idx, key in enumerate(bl_list, start=2):
        ws.cell(row=idx, column=3, value=key)
        ws.cell(row=idx, column=4, value=meta_datas.get(key))

    for idx, key in enumerate(fu_list, start=2):
        ws.cell(row=idx, column=5, value=key)
        ws.cell(row=idx, column=6, value=meta_datas.get(key))

    wb.save(os.path.abspath(os.path.join(REGISTRATION_DIR_127, "../..", "result.xlsx")))

    return None







