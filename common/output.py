
import os
import json
import SimpleITK as sitk

from tqdm import tqdm
from natsort import natsorted
from openpyxl import Workbook
from matplotlib import pyplot as plt

from medical.utils import get_meta_data
from medical.dicom_tags import METADATA_TAGS


def write_metadata_2_excel(dataset_dir: str) -> None:
    """
    Write the metadata to excel.
    :param dataset_dir: The directory of the dataset.
    :return: None
    """
    columns = ["CaseName", "PatientName", "PatientID", "StudyDate", "StudyTime", "Manufacturer", "Modality"]
    case_list = natsorted(os.listdir(os.path.join(dataset_dir)))
    meta_data = os.path.abspath(os.path.join(dataset_dir, "../", f"{dataset_dir.split(os.sep)[-1]}_metadata.json"))
    with open(meta_data, "r+") as f:
        meta_datas = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "MetaData"
    for idx, column in enumerate(columns, start=1):
        ws.cell(row=1, column=idx, value=column)
        ws.cell(row=1, column=idx+7, value=column)
    # 依照 case_list 的順序寫入
    for idx, case in enumerate(case_list, start=2):
        ws.cell(row=idx, column=1, value=f"{case}-1")
        for i, value in enumerate(meta_datas[f"{case}-1"]):
            ws.cell(row=idx, column=i+2, value=value)
        ws.cell(row=idx, column=1+7, value=f"{case}-2")
        for i, value in enumerate(meta_datas[f"{case}-2"]):
            ws.cell(row=idx, column=i+9, value=value)

    wb.save(os.path.abspath(os.path.join(dataset_dir, "..", f"{dataset_dir.split(os.sep)[-1]}_metadatas_result.xlsx")))
    return None


def write_metric_value_2_excel(dataset_dir: str) -> None:
    """
    Write the metric values to excel.
    :return: None
    """
    columns = ["Case name", "Correlation Metric Value", "Baseline", "Follow-up"]
    metric_value = os.path.abspath(os.path.join(dataset_dir, "..", f"{dataset_dir.split(os.sep)[-1]}_metric_value.json"))

    with open(metric_value, "r+") as f:
        metric_values = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "Result"
    for idx, column in enumerate(columns, start=1):
        ws.cell(row=1, column=idx, value=column)

    for idx, (key, value) in enumerate(metric_values.items(), start=2):
        ws.cell(row=idx, column=1, value=key)
        ws.cell(row=idx, column=2, value=value)

    wb.save(os.path.abspath(os.path.join(dataset_dir, "..", f"{dataset_dir.split(os.sep)[-1]}_metrics_result.xlsx")))

    return None


def write_metadata_2_json(datasets_dir: str) -> None:
    """
    Write the metadata to json.
    :param datasets_dir: The directory of the datasets.
    :return: None
    """
    case_list = natsorted(os.listdir(os.path.join(datasets_dir)))
    meta_datas = {}
    json_file = os.path.abspath(os.path.join(datasets_dir, "..", f"{datasets_dir.split(os.sep)[-1]}_metadata.json"))
    with tqdm(total=len(case_list)) as pbar:
        for case in case_list:
            pbar.set_description(f"{case} is getting metadata...")
            # if case in no_mask_list:
            #     pbar.update()
            #     continue
            dicom_bl_dir = os.path.join(datasets_dir, case, "dcm", "BL")
            dicom_fu_dir = os.path.join(datasets_dir, case, "dcm", "FU")
            dicom_bl = os.path.join(dicom_bl_dir, os.listdir(dicom_bl_dir)[0])
            dicom_fu = os.path.join(dicom_fu_dir, os.listdir(dicom_fu_dir)[0])
            bl_image = sitk.ReadImage(dicom_bl, sitk.sitkFloat32)
            fu_image = sitk.ReadImage(dicom_fu, sitk.sitkFloat32)
            for bl_fu, image in [("1", bl_image), ("2", fu_image)]:
                meta_datas[f"{case}-{bl_fu}"] = [
                    image.GetMetaData(METADATA_TAGS["PATIENT_NAME"]),
                    image.GetMetaData(METADATA_TAGS["PATIENT_ID"]),
                    image.GetMetaData(METADATA_TAGS["STUDY_DATE"]),
                    image.GetMetaData(METADATA_TAGS["STUDY_TIME"]).split(".")[0],
                    image.GetMetaData(METADATA_TAGS["MANUFACTURER"]),
                    image.GetMetaData(METADATA_TAGS["MODALITY"]),
                ]
            pbar.update()
    with open(json_file, "w") as f:
        json.dump(meta_datas, f)


def visualization_hist(json_file: str) -> None:
    """
    Visualize the histogram.
    :param json_file: The json file.
    :return: None
    """
    values = []
    with open(json_file, "r+") as f:
        file = json.load(f)
        for value in file.values():
            values.append(abs(value))
        plt.hist(values, bins=20, color="steelblue", edgecolor="k", alpha=0.65, rwidth=0.8)
        plt.xlabel("Registration metric value")
        plt.ylabel("Frequency")
        plt.title(f"Registration metric value histogram, total={len(values)}")
        plt.savefig(os.path.abspath(os.path.join(json_file, "..", f"{json_file.split(os.sep)[-1].split('metric_value')[0]}histogram.png")))
        # plt.show()

    return None
