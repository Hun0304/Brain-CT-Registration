"""
Excel input process class.
"""

import os

from tqdm.rich import tqdm
from natsort import natsorted
from openpyxl import load_workbook


class DataInputFromExcelProcessorBase:
    """
    The class of data input from Excel process.
    """

    def __init__(self, data_file: str, input_file_prefix: str):
        """
        The constructor of the class.
        :param data_file: The directory of the Excel file.
        :param input_file_prefix: The prefix of the input file.
        """
        self.__data_file = data_file
        self.__input_file_prefix = input_file_prefix

    def read_from_excel(self, sheet_name) -> None:
        """
        Read the data from Excel.
        :param sheet_name:
        :return: None
        """
        input_file = os.path.abspath(os.path.join(self.__data_file,
                                                  "..",
                                                  f"{self.__input_file_prefix}_{sheet_name}.xlsx"))
        wb = load_workbook(input_file)
        ws = wb.active
        header_mapping = {}
        for index, column in enumerate(ws[1], start=1):
            header_mapping[column.value] = index

        return None
