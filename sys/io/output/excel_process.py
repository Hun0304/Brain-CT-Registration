"""
Excel output process class.
"""
import os
import json
import SimpleITK as sitk

from typing import List, Dict
from tqdm.rich import tqdm
from natsort import natsorted
from openpyxl import Workbook

from medical.dicom_tags import MetaDataTags
from medical.utils import Calculator


class DataOutputToExcelProcessorBase:
    """
    The class of data output to Excel process.
    """

    def __init__(self, datasets_dir: str, output_file_prefix: str):
        """
        The constructor of the class.
        :param datasets_dir: The directory of the datasets.
        :param output_file_prefix: The prefix of the output file.
        """
        self.__datasets_dir = datasets_dir
        self.__output_file_prefix = output_file_prefix
        self.__case_list = natsorted(os.listdir(self.__datasets_dir))

    def write_to_excel(self, columns, data, sheet_name) -> None:
        """
        Write the data to Excel.
        :param columns:
        :param data:
        :param sheet_name:
        :return: None
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # ws.cell(row=1, column=1, value=columns[0])

        for idx, column in enumerate(columns, start=1):
            ws.cell(row=1, column=idx, value=column)

        for idx, case_data in enumerate(data, start=2):
            for i, value in enumerate(case_data):
                ws.cell(row=idx, column=i + 1, value=value)
        output_file = os.path.abspath(os.path.join(self.__datasets_dir,
                                                   "..",
                                                   f"{self.__output_file_prefix}_{sheet_name}.xlsx"))
        wb.save(output_file)

        return None


class HematomaExpansionDataOutputProcessor(DataOutputToExcelProcessorBase):
    """
    The class of Hematoma Expansion data output to Excel process.
    """

    def __init__(self, datasets_dir: str):
        """
        The constructor of the class.
        :param datasets_dir: The directory of the datasets.
        """
        super().__init__(datasets_dir, "HE_result")

    @staticmethod
    def generate_he_data(calc: Calculator, case_list: List, mask_bl_list: List, mask_fu_list: List) -> None:
        """
        Generate Hematoma Volume & Hematoma Expansion data.
        :param: calc:
        :param: case_list:
        :param: mask_bl_list:
        :param: mask_fu_list:
        :return: Yields case data
        """
        with tqdm(total=len(case_list)) as pbar:
            for mask_bl, mask_fu in zip(mask_bl_list, mask_fu_list):
                case_name = os.path.basename(mask_bl)[:os.path.basename(mask_bl).find('-')]
                pbar.set_description(f"{case_name} is calculating...")
                case_data = list(calc.calc_he(mask_bl, mask_fu))
                case_data.insert(0, case_name)
                yield case_data
                pbar.update()

    def write_he_to_excel(self) -> None:
        """
        Write the Hematoma Volume & Hematoma Expansion to Excel.
        :return: None
        """
        case_list = self.__case_list
        columns = ["Case number", "Baseline volume", "Follow volume", "Volume change(cc)", "Volume change ratio", "HE"]
        calc = Calculator()
        mask_bl_list = [os.path.join(self.__datasets_dir, case, "mask", "result", f"{case}-1-mask.nii.gz") for case in
                        case_list]
        mask_fu_list = [os.path.join(self.__datasets_dir, case, "mask", "result", f"{case}-2-mask.nii.gz") for case in
                        case_list]
        self.write_to_excel(columns, self.generate_he_data(calc, case_list, mask_bl_list, mask_fu_list), "Hematoma Expansion")

        return None


class MetadataDataOutputProcessor(DataOutputToExcelProcessorBase):
    """
    The class of metadata data output to Excel process.
    """

    def __init__(self, datasets_dir: str):
        """
        The constructor of the class.
        :param datasets_dir: The directory of the datasets.
        """
        super().__init__(datasets_dir, "metadata_result")

    def write_metadata_to_json(self) -> None:
        """
        Write the metadata to json.
        :return: None
        """
        case_list = self.__case_list
        meta_datas = {}
        json_file = os.path.abspath(os.path.join(self.__datasets_dir,
                                                 "..",
                                                 f"{self.__datasets_dir.split(os.sep)[-1]}_metadata.json"))
        meta_data_tags = MetaDataTags()
        with tqdm(total=len(case_list)) as pbar:
            for case in case_list:
                pbar.set_description(f"{case} is getting metadata...")
                dicom_bl_dir = os.path.join(self.__datasets_dir, case, "dcm", "BL")
                dicom_fu_dir = os.path.join(self.__datasets_dir, case, "dcm", "FU")
                dicom_bl = os.path.join(dicom_bl_dir, os.listdir(dicom_bl_dir)[0])
                dicom_fu = os.path.join(dicom_fu_dir, os.listdir(dicom_fu_dir)[0])
                bl_image = sitk.ReadImage(dicom_bl, sitk.sitkFloat32)
                fu_image = sitk.ReadImage(dicom_fu, sitk.sitkFloat32)
                for bl_fu, image in [("1", bl_image), ("2", fu_image)]:
                    meta_datas[f"{case}-{bl_fu}"] = [
                        image.GetMetaData(meta_data_tags.PATIENT_NAME),
                        image.GetMetaData(meta_data_tags.PATIENT_ID),
                        image.GetMetaData(meta_data_tags.STUDY_DATE),
                        image.GetMetaData(meta_data_tags.STUDY_TIME).split(".")[0],
                        image.GetMetaData(meta_data_tags.MANUFACTURER),
                        image.GetMetaData(meta_data_tags.MODALITY),
                    ]
                pbar.update()
        with open(json_file, "w") as f:
            json.dump(meta_datas, f)

        return None

    @staticmethod
    def generate_metadata(case_list: List, meta_datas) -> None:
        """
        Generate Metadata.
        :param: case_list:
        :param: meta_datas:
        :return: None
        """
        with tqdm(total=len(case_list)) as pbar:
            for case in case_list:
                pbar.set_description(f"{case}-1 is writing to excel...")
                yield [f"{case}-1"] + meta_datas[f"{case}-1"]
                yield [f"{case}-2"] + meta_datas[f"{case}-2"]
                pbar.update()

    def write_metadata_to_excel(self):
        """
        Write the metadata to Excel.
        :return: None
        """
        case_list = self.__case_list
        columns = ["CaseName", "PatientName", "PatientID", "StudyDate", "StudyTime", "Manufacturer", "Modality"]
        meta_data_file = os.path.abspath(os.path.join(self.__datasets_dir,
                                                      "..",
                                                      f"{self.__datasets_dir.split(os.sep)[-1]}_metadata.json"))
        with open(meta_data_file, "r+") as f:
            meta_datas = json.load(f)
        self.write_to_excel(columns, self.generate_metadata(case_list, meta_datas), "MetaData")

        return None


class MetricValueDataOutputProcessor(DataOutputToExcelProcessorBase):
    """
    The class of metric value data output to Excel process.
    """

    def __init__(self, datasets_dir: str):
        """
        The constructor of the class.
        :param datasets_dir: The directory of the datasets.
        """
        super().__init__(datasets_dir, "metrics_result")

    @staticmethod
    def generate_metric_values(metric_values: Dict) -> None:
        """
        :param: case_list:
        :param: datasets_dir:
        :return:
        """
        for key, value in metric_values.items():
            yield [key, value]

    def write_metric_value_to_excel(self) -> None:
        """
        Write the metric values to Excel.
        :return: None
        """
        columns = ["Case name", "Correlation Metric Value", "Baseline", "Follow-up"]
        metric_value = os.path.abspath(os.path.join(self.__datasets_dir,
                                                    "..",
                                                    f"{self.__datasets_dir.split(os.sep)[-1]}_metric_value.json"))

        with open(metric_value, "r+") as f:
            metric_values = json.load(f)

        self.write_to_excel(columns, self.generate_metric_values(metric_values), "Result")

        return None
