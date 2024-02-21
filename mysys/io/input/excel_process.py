"""
Excel input process class.
"""

import os

from typing import Tuple, Any, Generator
from collections import namedtuple, Counter

import openpyxl
from tqdm.rich import tqdm
from openpyxl import load_workbook


class DataInputFromExcelProcessorBase:
    """
    The class of data input from Excel process.
    """

    def __init__(self, datasets_dir: str, datasets_name: str):
        """
        The constructor of the class.
        :param datasets_dir: The directory of the datasets.
        :param datasets_name: The name of the datasets.
        """
        self.__datasets_dir = datasets_dir
        self.__datasets_name = datasets_name

    @staticmethod
    def generate_row_data(ws: openpyxl) -> Generator[Tuple[Any, Any], Any, None]:
        """
        Generate row data from Excel
        """
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row-1):
            yield row

    def readline_from_excel(self, sheet_name) -> Generator[Tuple[Any, Any], Any, None]:
        """
        Read the data from Excel.
        :param sheet_name:
        :return: None
        """
        input_file = os.path.abspath(os.path.join(self.__datasets_dir,
                                                  "..",
                                                  "output",
                                                  f"{self.__datasets_name}_{sheet_name}.xlsx"))

        if not os.path.isfile(input_file):
            raise FileExistsError(f"{input_file} is not exists.")

        wb = load_workbook(input_file)
        ws = wb.active
        wb.close()

        return self.generate_row_data(ws)


class HematomaExpansionDataInputProcessor(DataInputFromExcelProcessorBase):
    """
    The class of Hematoma Expansion data input from Excel process.
    """

    def __init__(self, datasets_dir: str, datasets_name: str):
        """
        The constructor of the class.
        :param datasets_dir: The directory of the datasets.
        :param datasets_name: The name of the datasets.
        """
        super().__init__(datasets_dir, datasets_name)

    def read_he_data(self) -> Generator[Tuple[str, float], Any, None]:
        """
        Read Hematoma Expansion data from Excel.
        :return:
        """
        HE_Data = namedtuple("HE_Data", ["CaseNumber", "HE"])
        for row in self.readline_from_excel("HE_result"):
            CaseNumber = row[0].value
            HE = row[-1].value
            yield HE_Data(CaseNumber, HE)

        return None

    def he_data_reader(self) -> Generator[Tuple[str, float], Any, None]:
        """
        Read Hematoma Expansion data from Excel.
        :return:
        """
        counter = Counter()
        for case, he in self.read_he_data():
            if he == 1:
                counter.update(["HE"])
            yield case, he, counter

        return None
